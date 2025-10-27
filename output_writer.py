import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from pathlib import Path
from datetime import datetime
import webbrowser

def color_for_signal(signal):
    s = signal.upper()
    if "STRONG BUY" in s: return "#00b050"
    if "BUY" in s: return "#92d050"
    if "STRONG WAIT" in s: return "#ff0000"
    if "WAIT" in s: return "#ff9999"
    return "#ffffff"

def confidence_bar(conf):
    pct = int(conf.replace("%", ""))
    color = "#00b050" if pct > 75 else "#ffd966" if pct >= 50 else "#ff6666"
    return f"<div style='width:100px;background:#eee;border-radius:4px;'><div style='width:{pct}%;background:{color};height:12px;border-radius:4px;'></div></div>"

def write_outputs(df):
    excel_path, html_path = Path("ETF_Screener_Pro_2026.xlsx"), Path("ETF_Screener_Pro_2026.html")
    df.to_excel(excel_path, index=False)

    # Excel color coding
    wb = load_workbook(excel_path)
    ws = wb.active
    for r in range(2, ws.max_row + 1):
        fill = color_for_signal(str(ws[f"I{r}"].value))
        ws[f"I{r}"].fill = PatternFill(start_color=fill.replace("#", ""), end_color=fill.replace("#", ""), fill_type="solid")
    wb.save(excel_path)

    # Prepare HTML data
    html_df = df.copy()
    html_df["Signal"] = html_df["Signal"].apply(lambda s: f"<b style='color:white;background:{color_for_signal(s)};padding:4px 8px;border-radius:4px;'>{s}</b>")
    html_df["Confidence"] = html_df["Confidence"].apply(confidence_bar)
    cols = ["Ticker", "Sparkline", "Chart", "Close", "RSI", "50DMA", "200DMA", "ZScore", "Signal", "Score", "Confidence"]
    html_df = html_df[cols]

    # Get current local timestamp
    last_updated = datetime.now().strftime("%d %B %Y, %I:%M %p")

    # HTML dashboard
    styled_html = f"""
    <html>
    <head>
      <title>ETF Screener Dashboard</title>
      <style>
        body {{ font-family: Arial, sans-serif; background:#fafafa; }}
        .header {{ text-align:center; margin-top:30px; font-size:22px; font-weight:bold; color:#0047AB; }}
        .timestamp {{ text-align:center; color:#444; margin-bottom:20px; font-size:14px; }}
        table {{ border-collapse: collapse; width:95%; margin:30px auto; box-shadow:0 0 10px rgba(0,0,0,0.1); background:white; }}
        th, td {{ padding:10px 15px; text-align:center; border-bottom:1px solid #ddd; }}
        th {{ background:#0047AB; color:white; font-size:14px; }}
        tr:hover {{ background:#f1f7ff; }}
        caption {{ caption-side:top; font-weight:bold; font-size:18px; padding:10px; }}
      </style>
    </head>
    <body>
      <div class="header">ETF Screener Pro 2026 Dashboard</div>
      <div class="timestamp">Last Updated: {last_updated}</div>
      <div style="margin-top:20px;">
        <a href="portfolio_dashboard.html" target="_blank"
            style="background:#007bff; color:white; padding:10px 20px; border-radius:8px; text-decoration:none;">
            📈 View Portfolio Tracker
        </a>
      </div>
      <table>
        <caption>Sorted by Z-Score</caption>
        {html_df.to_html(index=False, escape=False)}
      </table>
    </body>
    </html>
    """

    # Write and open
    Path(html_path).write_text(styled_html, encoding="utf-8")
    print(f"✅ Excel & HTML dashboard generated:\n  📊 {excel_path}\n  🌐 {html_path}")
    print(f"path to open : {str(html_path.resolve())}")
    # Auto-open in browser
    webbrowser.open_new_tab(str(html_path.resolve()))

    return str(html_path)
